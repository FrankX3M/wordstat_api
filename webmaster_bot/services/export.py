
import csv
import json
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Callable, Optional
import asyncio

from config import EXPORTS_DIR, MAX_EXPORT_ROWS, DEFAULT_PAGE_SIZE
from services.api import YandexWebmasterAPI
from utils.logger import setup_logger, log_exception

logger = setup_logger(__name__)


class ExportService:
    """Сервис для экспорта данных"""
    
    def __init__(self, api: YandexWebmasterAPI):
        self.api = api
    
    async def create_export(
        self,
        host_id: str,
        export_type: str,
        device_type: str,
        date_from: str,
        date_to: str,
        export_format: str = "csv",
        progress_callback: Optional[Callable] = None
    ) -> str:
        """Создание экспорта"""
        
        logger.info(f"Creating export: type={export_type}, device={device_type}, format={export_format}")
        logger.info(f"Date range: {date_from} to {date_to}")
        
        # Сбор данных
        if export_type == "popular":
            data = await self._export_popular_queries(
                host_id, date_from, date_to, device_type, progress_callback
            )
        elif export_type == "history":
            data = await self._export_history(
                host_id, date_from, date_to, device_type, progress_callback
            )
        elif export_type == "history_all":
            data = await self._export_history_all(
                host_id, date_from, date_to, device_type, progress_callback
            )
        elif export_type == "analytics":
            data = await self._export_analytics(
                host_id, date_from, date_to, device_type, progress_callback
            )
        elif export_type == "enhanced":
            data = await self._export_enhanced(
                host_id, date_from, date_to, device_type, progress_callback
            )
        else:
            raise ValueError(f"Unknown export type: {export_type}")
        
        # Создание файла
        filename = self._generate_filename(host_id, export_type, export_format)
        file_path = Path(EXPORTS_DIR) / filename
        
        if export_format == "csv":
            self._save_as_csv(data, file_path)
        elif export_format == "xlsx":
            self._save_as_xlsx(data, file_path)
        elif export_format == "json":
            self._save_as_json(data, file_path)
        else:
            raise ValueError(f"Unknown export format: {export_format}")
        
        logger.info(f"Export saved: {file_path}")
        return str(file_path)
    
    def _extract_indicators(self, query: Dict) -> Dict:
        """
        Извлечение индикаторов из различных возможных структур ответа API
        """
        result = {}
        
        # Вариант 1: indicators как объект
        if "indicators" in query and isinstance(query["indicators"], dict):
            indicators = query["indicators"]
            logger.debug(f"Found indicators dict: {indicators}")
            
            # Прямые значения
            for key in ["TOTAL_SHOWS", "TOTAL_CLICKS", "AVG_SHOW_POSITION", 
                       "AVG_CLICK_POSITION", "CTR", "DEMAND"]:
                if key in indicators:
                    result[key] = indicators[key]
            
            # Вложенные объекты (если есть)
            for key, value in indicators.items():
                if isinstance(value, dict):
                    for sub_key, sub_value in value.items():
                        result[f"{key}_{sub_key}"] = sub_value
        
        # Вариант 2: indicators как массив
        elif "indicators" in query and isinstance(query["indicators"], list):
            indicators_list = query["indicators"]
            logger.debug(f"Found indicators list with {len(indicators_list)} items")
            
            if indicators_list:
                # Берем последний элемент (самые свежие данные)
                latest = indicators_list[-1]
                if isinstance(latest, dict):
                    for key, value in latest.items():
                        if key != "date":
                            result[key] = value
        
        # Вариант 3: данные на верхнем уровне
        else:
            logger.debug("No indicators field, checking top level")
            for key in ["TOTAL_SHOWS", "TOTAL_CLICKS", "AVG_SHOW_POSITION",
                       "AVG_CLICK_POSITION", "CTR", "DEMAND",
                       "total_shows", "total_clicks", "avg_show_position",
                       "avg_click_position", "ctr"]:
                if key in query:
                    result[key.upper()] = query[key]
        
        # Логируем что нашли
        if result:
            logger.debug(f"Extracted indicators: {result}")
        else:
            logger.warning(f"No indicators extracted from query: {query.get('query_text', 'unknown')}")
            # Логируем полную структуру для отладки
            logger.debug(f"Full query structure: {json.dumps(query, ensure_ascii=False)[:500]}")
        
        return result
    
    async def _export_popular_queries(
        self,
        host_id: str,
        date_from: str,
        date_to: str,
        device_type: str,
        progress_callback: Optional[Callable] = None
    ) -> List[Dict]:
        """Экспорт популярных запросов с детальной диагностикой"""
        
        all_queries = []
        offset = 0
        page_size = min(DEFAULT_PAGE_SIZE, 500)
        
        logger.info(f"Starting popular queries export for host {host_id}")
        logger.info(f"Parameters: date_from={date_from}, date_to={date_to}, device={device_type}")
        
        while len(all_queries) < MAX_EXPORT_ROWS:
            try:
                logger.info(f"Fetching page at offset {offset}, limit {page_size}")
                
                result = await self.api.get_search_queries(
                    host_id=host_id,
                    date_from=date_from,
                    date_to=date_to,
                    device_type=device_type,
                    limit=page_size,
                    offset=offset,
                    order_by="TOTAL_SHOWS"
                )
                
                # Детальное логирование ответа API
                logger.info(f"API Response keys: {result.keys()}")
                logger.info(f"Total count from API: {result.get('count', 0)}")
                
                queries = result.get("queries", [])
                logger.info(f"Got {len(queries)} queries in this page")
                
                if not queries:
                    logger.info("No more queries, breaking")
                    break
                
                # Логируем структуру первого запроса для отладки
                if queries and offset == 0:
                    first_query = queries[0]
                    logger.info("=" * 60)
                    logger.info("FIRST QUERY STRUCTURE:")
                    logger.info(json.dumps(first_query, ensure_ascii=False, indent=2))
                    logger.info("=" * 60)
                
                # Обработка каждого запроса
                for idx, query in enumerate(queries):
                    row = {
                        "query_id": query.get("query_id", ""),
                        "query_text": query.get("query_text", ""),
                    }
                    
                    # Извлекаем индикаторы
                    indicators = self._extract_indicators(query)
                    
                    if indicators:
                        row.update(indicators)
                    else:
                        # Если индикаторов нет, пробуем альтернативные поля
                        logger.warning(f"No indicators for query: {query.get('query_text')}")
                        
                        # Добавляем пустые поля
                        row.update({
                            "TOTAL_SHOWS": query.get("total_shows", 0),
                            "TOTAL_CLICKS": query.get("total_clicks", 0),
                            "AVG_SHOW_POSITION": query.get("avg_show_position", 0),
                            "AVG_CLICK_POSITION": query.get("avg_click_position", 0),
                            "CTR": query.get("ctr", 0)
                        })
                    
                    all_queries.append(row)
                
                offset += len(queries)
                
                # Обновление прогресса
                if progress_callback:
                    total_found = result.get("count", len(all_queries))
                    await progress_callback(
                        len(all_queries),
                        min(total_found, MAX_EXPORT_ROWS),
                        f"Загружено запросов: {len(all_queries)}"
                    )
                
                # Если получили все данные
                if len(queries) < page_size:
                    logger.info("Received less than page_size, all data fetched")
                    break
                
            except Exception as e:
                logger.error(f"Error fetching popular queries at offset {offset}")
                log_exception(logger, e, "_export_popular_queries")
                break
        
        logger.info(f"✅ Exported {len(all_queries)} popular queries")
        
        # Финальная проверка данных
        non_zero_count = sum(1 for q in all_queries if any(
            q.get(k, 0) != 0 for k in ["TOTAL_SHOWS", "TOTAL_CLICKS"]
        ))
        logger.info(f"Queries with non-zero data: {non_zero_count}/{len(all_queries)}")
        
        return all_queries[:MAX_EXPORT_ROWS]
    
    async def _export_history(
        self,
        host_id: str,
        date_from: str,
        date_to: str,
        device_type: str,
        progress_callback: Optional[Callable] = None
    ) -> List[Dict]:
        """Экспорт истории запросов"""
        
        logger.info("Starting history export...")
        
        # Сначала получаем топ-100 популярных запросов
        result = await self.api.get_search_queries(
            host_id=host_id,
            date_from=date_from,
            date_to=date_to,
            device_type=device_type,
            limit=100,
            offset=0,
            order_by="TOTAL_SHOWS"
        )
        
        top_queries = result.get("queries", [])
        logger.info(f"Got {len(top_queries)} top queries for history")
        
        # Для каждого запроса получаем историю
        history_data = []
        total_queries = len(top_queries)
        
        for idx, query_data in enumerate(top_queries, 1):
            try:
                query_text = query_data.get("query_text")
                query_id = query_data.get("query_id")
                
                if not query_text:
                    continue
                
                logger.debug(f"Fetching history for query: {query_text}")
                
                # Получаем историю для этого запроса
                history = await self.api.get_search_queries_history(
                    host_id=host_id,
                    query_indicator=query_text,
                    date_from=date_from,
                    date_to=date_to,
                    device_type=device_type
                )
                
                # Логируем структуру истории для первого запроса
                if idx == 1:
                    logger.info("=" * 60)
                    logger.info("HISTORY STRUCTURE:")
                    logger.info(json.dumps(history, ensure_ascii=False, indent=2)[:1000])
                    logger.info("=" * 60)
                
                # Обработка точек истории
                indicators_list = history.get("indicators", [])
                logger.debug(f"Got {len(indicators_list)} history points for {query_text}")
                
                for point in indicators_list:
                    row = {
                        "query_id": query_id,
                        "query_text": query_text,
                        "date": point.get("date", ""),
                    }
                    
                    # Извлекаем indicators для каждой даты
                    point_indicators = self._extract_indicators(point)
                    if point_indicators:
                        row.update(point_indicators)
                    
                    history_data.append(row)
                
                # Обновление прогресса
                if progress_callback:
                    await progress_callback(
                        idx,
                        total_queries,
                        f"Обработано запросов: {idx}/{total_queries}"
                    )
                
                # Задержка между запросами
                await asyncio.sleep(0.1)
                
            except Exception as e:
                logger.warning(f"Error fetching history for query '{query_text}': {e}")
                continue
        
        logger.info(f"✅ Exported {len(history_data)} history points")
        return history_data
    
    async def _export_history_all(
        self,
        host_id: str,
        date_from: str,
        date_to: str,
        device_type: str,
        progress_callback: Optional[Callable] = None
    ) -> List[Dict]:
        """Расширенная история запросов"""
        
        # Используем тот же подход что и в history, но с большим количеством запросов
        logger.info("Starting history_all export...")
        
        result = await self.api.get_search_queries(
            host_id=host_id,
            date_from=date_from,
            date_to=date_to,
            device_type=device_type,
            limit=500,
            offset=0,
            order_by="TOTAL_SHOWS"
        )
        
        queries = result.get("queries", [])
        all_history = []
        
        # Ограничиваем до 200 запросов для производительности
        for idx, query_data in enumerate(queries[:200], 1):
            try:
                query_text = query_data.get("query_text")
                query_id = query_data.get("query_id")
                
                if not query_text:
                    continue
                
                history = await self.api.get_search_queries_history(
                    host_id=host_id,
                    query_indicator=query_text,
                    date_from=date_from,
                    date_to=date_to,
                    device_type=device_type
                )
                
                indicators_list = history.get("indicators", [])
                
                for point in indicators_list:
                    row = {
                        "query_id": query_id,
                        "query_text": query_text,
                        "date": point.get("date", ""),
                    }
                    
                    point_indicators = self._extract_indicators(point)
                    if point_indicators:
                        row.update(point_indicators)
                    
                    all_history.append(row)
                
                if progress_callback and idx % 10 == 0:
                    await progress_callback(
                        idx,
                        min(len(queries), 200),
                        f"Обработано: {idx}"
                    )
                
                await asyncio.sleep(0.05)
                
            except Exception as e:
                logger.warning(f"Error in history_all: {e}")
                continue
        
        logger.info(f"✅ Exported {len(all_history)} history_all records")
        return all_history
    
    async def _export_analytics(
        self,
        host_id: str,
        date_from: str,
        date_to: str,
        device_type: str,
        progress_callback: Optional[Callable] = None
    ) -> List[Dict]:
        """Детальная аналитика"""
        
        logger.info("Starting analytics export...")
        
        popular_result = await self.api.get_search_queries(
            host_id=host_id,
            date_from=date_from,
            date_to=date_to,
            device_type=device_type,
            limit=200,
            offset=0,
            order_by="TOTAL_SHOWS"
        )
        
        queries = popular_result.get("queries", [])
        analytics_data = []
        
        for idx, query_data in enumerate(queries, 1):
            try:
                query_text = query_data.get("query_text")
                query_id = query_data.get("query_id")
                
                if not query_text:
                    continue
                
                row = {
                    "query_id": query_id,
                    "query_text": query_text,
                }
                
                # Извлекаем индикаторы
                indicators = self._extract_indicators(query_data)
                if indicators:
                    row.update(indicators)
                
                # Получаем историю для трендов
                try:
                    history = await self.api.get_search_queries_history(
                        host_id=host_id,
                        query_indicator=query_text,
                        date_from=date_from,
                        date_to=date_to,
                        device_type=device_type
                    )
                    
                    indicators_list = history.get("indicators", [])
                    
                    if len(indicators_list) >= 2:
                        row["history_points"] = len(indicators_list)
                        
                        first_point = self._extract_indicators(indicators_list[0])
                        last_point = self._extract_indicators(indicators_list[-1])
                        
                        # Тренд показов
                        first_shows = first_point.get("TOTAL_SHOWS", 0)
                        last_shows = last_point.get("TOTAL_SHOWS", 0)
                        
                        if first_shows > 0:
                            trend = ((last_shows - first_shows) / first_shows) * 100
                            row["shows_trend_percent"] = round(trend, 2)
                        
                        # Тренд кликов
                        first_clicks = first_point.get("TOTAL_CLICKS", 0)
                        last_clicks = last_point.get("TOTAL_CLICKS", 0)
                        
                        if first_clicks > 0:
                            trend = ((last_clicks - first_clicks) / first_clicks) * 100
                            row["clicks_trend_percent"] = round(trend, 2)
                
                except Exception as e:
                    logger.debug(f"Could not get history for analytics: {e}")
                
                analytics_data.append(row)
                
                if progress_callback and idx % 10 == 0:
                    await progress_callback(
                        idx,
                        len(queries),
                        f"Обработано: {idx}/{len(queries)}"
                    )
                
                await asyncio.sleep(0.05)
                
            except Exception as e:
                logger.warning(f"Error in analytics: {e}")
                continue
        
        logger.info(f"✅ Exported {len(analytics_data)} analytics records")
        return analytics_data
    
    async def _export_enhanced(
        self,
        host_id: str,
        date_from: str,
        date_to: str,
        device_type: str,
        progress_callback: Optional[Callable] = None
    ) -> List[Dict]:
        """Расширенный экспорт"""
        
        logger.info("Starting enhanced export...")
        
        # Получаем максимум популярных запросов
        all_queries = []
        offset = 0
        page_size = 500
        
        while len(all_queries) < min(MAX_EXPORT_ROWS, 1000):
            try:
                result = await self.api.get_search_queries(
                    host_id=host_id,
                    date_from=date_from,
                    date_to=date_to,
                    device_type=device_type,
                    limit=page_size,
                    offset=offset,
                    order_by="TOTAL_SHOWS"
                )
                
                queries = result.get("queries", [])
                
                if not queries:
                    break
                
                all_queries.extend(queries)
                offset += len(queries)
                
                if len(queries) < page_size:
                    break
                
            except Exception as e:
                logger.error(f"Error fetching queries at offset {offset}")
                break
        
        logger.info(f"Got {len(all_queries)} queries for enhanced export")
        
        # Обрабатываем каждый запрос
        enhanced_data = []
        
        for idx, query_data in enumerate(all_queries, 1):
            try:
                query_text = query_data.get("query_text")
                query_id = query_data.get("query_id")
                
                if not query_text:
                    continue
                
                row = {
                    "query_id": query_id,
                    "query_text": query_text,
                    "device_type": device_type,
                    "period_from": date_from,
                    "period_to": date_to,
                }
                
                # Извлекаем индикаторы
                indicators = self._extract_indicators(query_data)
                if indicators:
                    row.update(indicators)
                
                enhanced_data.append(row)
                
                if progress_callback and idx % 50 == 0:
                    await progress_callback(
                        idx,
                        len(all_queries),
                        f"Обработано: {idx}/{len(all_queries)}"
                    )
                
            except Exception as e:
                logger.warning(f"Error in enhanced export: {e}")
                continue
        
        logger.info(f"✅ Exported {len(enhanced_data)} enhanced records")
        return enhanced_data
    
    def _generate_filename(self, host_id: str, export_type: str, export_format: str) -> str:
        """Генерация имени файла"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_host_id = host_id.replace(":", "_").replace("/", "_")[:50]
        return f"export_{safe_host_id}_{export_type}_{timestamp}.{export_format}"
    
    def _save_as_csv(self, data: List[Dict], file_path: Path):
        """Сохранение в CSV"""
        
        if not data:
            with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(["No data available"])
            return
        
        all_keys = set()
        for item in data:
            all_keys.update(item.keys())
        
        priority_keys = ["query_id", "query_text", "date", "device_type", "period_from", "period_to",
                        "TOTAL_SHOWS", "TOTAL_CLICKS", "CTR", "AVG_SHOW_POSITION", "AVG_CLICK_POSITION"]
        fieldnames = [k for k in priority_keys if k in all_keys]
        fieldnames.extend(sorted([k for k in all_keys if k not in priority_keys]))
        
        with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(data)
        
        logger.info(f"Saved {len(data)} rows to CSV with {len(fieldnames)} columns")
    
    def _save_as_xlsx(self, data: List[Dict], file_path: Path):
        """Сохранение в Excel"""
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Export"
            
            if not data:
                ws['A1'] = "No data available"
                wb.save(file_path)
                return
            
            all_keys = set()
            for item in data:
                all_keys.update(item.keys())
            
            priority_keys = ["query_id", "query_text", "date", "device_type", "period_from", "period_to",
                            "TOTAL_SHOWS", "TOTAL_CLICKS", "CTR", "AVG_SHOW_POSITION", "AVG_CLICK_POSITION"]
            headers = [k for k in priority_keys if k in all_keys]
            headers.extend(sorted([k for k in all_keys if k not in priority_keys]))
            
            # Заголовки
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Данные
            for row_idx, item in enumerate(data, 2):
                for col_idx, key in enumerate(headers, 1):
                    value = item.get(key)
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    if isinstance(value, (int, float)) and not isinstance(value, bool):
                        if key in ["CTR", "ctr", "clicks_trend_percent", "shows_trend_percent"]:
                            cell.number_format = '0.00'
                        elif isinstance(value, float):
                            cell.number_format = '0.00'
                        else:
                            cell.number_format = '#,##0'
            
            # Автоширина
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            ws.freeze_panes = "A2"
            wb.save(file_path)
            logger.info(f"Saved {len(data)} rows to Excel")
            
        except ImportError:
            logger.warning("openpyxl not available, falling back to CSV")
            self._save_as_csv(data, file_path.with_suffix('.csv'))
    
    def _save_as_json(self, data: List[Dict], file_path: Path):
        """Сохранение в JSON"""
        
        output = {
            "export_date": datetime.now().isoformat(),
            "total_records": len(data),
            "data": data
        }
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(output, f, ensure_ascii=False, indent=2)
        
        logger.info(f"Saved {len(data)} items to JSON")

